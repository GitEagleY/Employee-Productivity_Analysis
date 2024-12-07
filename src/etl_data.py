from openpyxl import load_workbook
import pandas as pd
# Load raw data
wb = load_workbook("../Data/employee_tasks_raw.xlsx")
sheet = wb.active

# Create new columns for ETL
sheet["H1"] = "DelayDays"
sheet["I1"] = "CompletionPercentage"

# Add delay calculation
for row in range(2, sheet.max_row + 1):
    deadline = sheet[f"E{row}"].value
    completion_date = sheet[f"F{row}"].value

    if completion_date:
        delay = (completion_date - deadline).days
        sheet[f"H{row}"] = delay if delay > 0 else 0
        sheet[f"I{row}"] = 100
    else:
        sheet[f"H{row}"] = None
        sheet[f"I{row}"] = 0

# Save transformed data
wb.save("../Data/employee_tasks_transformed.xlsx")
print("Transformed data saved: employee_tasks_transformed.xlsx")